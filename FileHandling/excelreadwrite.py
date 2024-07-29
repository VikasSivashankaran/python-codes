import openpyxl
from openpyxl import Workbook

def main():
    print("Script Started")
    
    workbook = openpyxl.load_workbook('data1.xlsx')
    sheet = workbook.active
    
    data = []

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        row_data = list(row)
        
        if i == 0:
            row_data.append('Final Amount')
            row_data.append('Amount after 50 percentage discount')
        else:
            # Check if any of the required cells are None
            if row[2] is None or row[3] is None or row[4] is None:
                print(f"Skipping row {i} due to None value: {row}")
                row_data.append(None)
                row_data.append(None)
            else:
                amount = int(row[2])
                qty = int(row[3])
                gst = int(row[4])
                final_amount = amount * qty
                gst_amount = final_amount * (gst / 100)
                final_total = final_amount + gst_amount
                row_data.append(final_total)

                percentage_of_50 = final_total * 0.5
                row_data.append(percentage_of_50)
        
        data.append(row_data)

    # Create a new workbook and select the active worksheet
    output_workbook = Workbook()
    output_sheet = output_workbook.active

    # Write the data to the new workbook
    for row_data in data:
        output_sheet.append(row_data)
    
    # Save the new workbook
    output_workbook.save('data2.xlsx')

    print("Script Completed")

main()
